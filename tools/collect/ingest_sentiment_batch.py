# -*- coding: utf-8 -*-
"""
ingest_sentiment_batch.py — 批量导入「情感泛化词 + 品线」联合搜索的 VOC 数据。

与 ingest_competitor_zip.py 不同，本脚本：
  1. 支持多国家（从文件名自动解析 ISO 码）
  2. 支持多品线（吸奶器/喂养电器/家居出行）
  3. 支持多平台类型（reddit/tiktok/facebook/mumsnet/trustpilot 等）
  4. 处理目录下全部 .xlsx 文件（非 ZIP 解压）

文件命名规范：
  {platform}_{country_iso}_{product_line}_{type}_merged.xlsx
  例：reddit_us_breastpump_sentiment_merged.xlsx
      tiktok_uk_bottlewarmer_sentiment_merged.xlsx
      mumsnet_uk_stroller_sentiment_merged.xlsx

用法：
    python tools/collect/ingest_sentiment_batch.py [--dry-run] [--dir path/to/dir]
"""
from __future__ import annotations

import csv
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

PROJ = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(PROJ))

from tools.cleaning._common import (
    normalize_brand, infer_pain_subcategory, PAIN_SUBCATEGORY_PARENT,
    COUNTRY_CODE_MAP,
)

DEFAULT_BATCH_DIR = PROJ / "data" / "add_data" / "02_sentiment_batch"
TARGET_CSV = PROJ / "data" / "delivery" / "tables" / "dim_voc_negative_extract.csv"
STAGING_DIR = PROJ / "data" / "processed"

CSV_20_HEADERS = [
    "国家", "区域cluster", "产品品线", "平台类型", "平台",
    "画像编码", "画像名称", "生命周期", "痛点大类(功能/价格/体验/服务/安全)",
    "负面主题", "负面原文摘录(本地语言)", "负面原文摘录(中文翻译)",
    "频次估算", "负面强度(高/中/低)", "竞品关联品牌", "对应运营建议",
    "来源URL", "采集日期", "批次编码", "优先级",
]

# ── 文件名解析映射 ──

COUNTRY_ISO_TO_CN: dict[str, str] = {
    "us": "美国", "uk": "英国", "gb": "英国", "ca": "加拿大", "au": "澳大利亚",
    "de": "德国", "fr": "法国", "es": "西班牙", "mx": "墨西哥",
    "ae": "阿联酋", "sa": "沙特阿拉伯", "it": "意大利", "nl": "荷兰",
    "be": "比利时", "at": "奥地利", "se": "瑞典", "no": "挪威",
    "pl": "波兰", "jp": "日本", "kr": "韩国", "sg": "新加坡",
    "my": "马来西亚", "th": "泰国", "ph": "菲律宾", "vn": "越南",
    "in": "印度", "br": "巴西", "cn": "中国", "nz": "新西兰",
}

COUNTRY_CLUSTER: dict[str, str] = {
    "美国": "北美高购买力区", "加拿大": "北美高购买力区",
    "英国": "西欧高信任论坛区", "德国": "西欧高信任论坛区", "法国": "西欧高信任论坛区",
    "澳大利亚": "英语口碑圈", "新西兰": "英语口碑圈",
    "西班牙": "南欧拉美社媒区", "墨西哥": "南欧拉美社媒区",
    "阿联酋": "中东视觉社媒区", "沙特阿拉伯": "中东视觉社媒区",
    "意大利": "南欧拉美社媒区", "荷兰": "西欧高信任论坛区", "比利时": "西欧高信任论坛区",
}

PRODUCT_LINE_MAP: dict[str, str] = {
    "breastpump": "吸奶器", "pump": "吸奶器",
    "bottlewarmer": "喂养电器", "sterilizer": "喂养电器", "feeding": "喂养电器", "foodmaker": "喂养电器",
    "stroller": "家居出行", "carseat": "家居出行", "carrier": "家居出行", "pram": "家居出行",
}

PLATFORM_NAME_MAP: dict[str, str] = {
    "reddit": "Reddit", "tiktok": "TikTok", "instagram": "Instagram",
    "facebook": "Facebook", "mumsnet": "Mumsnet", "netmums": "Netmums",
    "whattoexpect": "What to Expect Forums", "babycenter": "BabyCenter",
    "trustpilot": "Trustpilot", "urbia": "Urbia.de",
    "magicmaman": "Magicmaman", "mumzworld": "Mumzworld",
}

PLATFORM_TYPE_MAP: dict[str, str] = {
    "reddit": "垂类社区类", "mumsnet": "垂类社区类", "netmums": "垂类社区类",
    "whattoexpect": "垂类社区类", "babycenter": "垂类社区类", "urbia": "垂类社区类",
    "magicmaman": "垂类社区类", "mumzworld": "垂类社区类",
    "tiktok": "社媒传播类", "instagram": "社媒传播类", "facebook": "社媒传播类",
    "trustpilot": "第三方评测类",
}

NEGATIVE_KEYWORDS = [
    "issue", "problem", "hate", "awful", "terrible", "horrible", "broken",
    "worst", "disappointed", "frustrat", "leak", "broke", "defect", "fail",
    "waste", "useless", "regret", "return", "refund", "pain", "hurt",
    "suck", "crap", "garbage", "annoying", "malfunction", "recall",
    "kaputt", "defekt", "schmerzhaft", "enttäuschend", "minderwertig",
    "cassé", "fuite", "douloureux", "déçu", "dangereux",
    "roto", "doloroso", "decepcionado", "peligroso", "defectuoso",
]


def parse_filename(fname: str) -> dict[str, str] | None:
    """Parse structured filename into metadata. Returns None if unparseable."""
    stem = Path(fname).stem.lower()
    parts = stem.replace("-", "_").split("_")
    if len(parts) < 3:
        return None

    platform_key = parts[0]
    country_iso = parts[1]
    product_key = parts[2]

    country_cn = COUNTRY_ISO_TO_CN.get(country_iso)
    product_line = PRODUCT_LINE_MAP.get(product_key)
    platform_name = PLATFORM_NAME_MAP.get(platform_key, platform_key.title())
    platform_type = PLATFORM_TYPE_MAP.get(platform_key, "垂类社区类")

    if not country_cn:
        print(f"  WARN: unknown country ISO '{country_iso}' in {fname}")
        country_cn = country_iso.upper()

    if not product_line:
        print(f"  WARN: unknown product line '{product_key}' in {fname}")
        product_line = "吸奶器"

    cluster = COUNTRY_CLUSTER.get(country_cn, "其他区域")

    return {
        "country": country_cn,
        "country_iso": country_iso.upper(),
        "cluster": cluster,
        "product_line": product_line,
        "platform": platform_name,
        "platform_type": platform_type,
    }


def _is_negative(emotion: str | None, post_type: str | None, text: str | None) -> bool:
    e = (emotion or "").strip()
    pt = (post_type or "").strip()
    if e in ("差评", "negative", "neg"):
        return True
    if pt in ("吐槽", "complaint") and e not in ("好评", "positive"):
        return True
    if e in ("好评", "positive"):
        return False
    if text:
        lower = text.lower()
        if sum(1 for kw in NEGATIVE_KEYWORDS if kw in lower) >= 2:
            return True
    return False


def _infer_pain_category(text: str, topics: str) -> str:
    combined = (text + " " + topics).lower()
    if any(w in combined for w in ["price", "expensive", "cost", "afford", "cheap", "贵", "价格", "teuer", "cher", "caro"]):
        return "价格"
    if any(w in combined for w in ["suction", "power", "motor", "function", "功能", "吸力", "Saugkraft", "aspiration"]):
        return "功能"
    if any(w in combined for w in ["service", "support", "customer", "warranty", "服务", "售后", "Kundendienst"]):
        return "服务"
    if any(w in combined for w in ["safe", "danger", "hazard", "recall", "安全", "危险", "unsicher", "dangereux", "peligroso"]):
        return "安全"
    return "体验"


def _infer_intensity(text: str) -> str:
    lower = text.lower()
    strong_words = [
        "hate", "worst", "terrible", "horrible", "awful", "waste of money",
        "never again", "dangerous", "defect", "recall", "scam",
        "kaputt", "gefährlich", "dangereux", "peligroso",
    ]
    if any(w in lower for w in strong_words):
        return "高"
    mild_words = ["okay", "not great", "could be better", "minor", "slight"]
    if any(w in lower for w in mild_words):
        return "低"
    return "中"


# ── Column detection: auto-detect text/emotion/brand columns ──

TEXT_COLUMN_CANDIDATES = ["统一分析文本", "text", "content", "review", "comment", "正文", "评论内容", "原文"]
EMOTION_COLUMN_CANDIDATES = ["情绪", "sentiment", "emotion", "rating", "评分"]
BRAND_COLUMN_CANDIDATES = ["提及品牌", "brand", "品牌", "competitor_brand"]
TOPIC_COLUMN_CANDIDATES = ["主要话题", "topic", "话题", "category"]
URL_COLUMN_CANDIDATES = ["来源URL", "url", "link", "source_url", "帖子ID"]
DATE_COLUMN_CANDIDATES = ["抓取时间", "date", "created_at", "采集日期", "时间"]


def _find_column(headers: list[str], candidates: list[str]) -> str | None:
    header_lower = {h.lower(): h for h in headers if h}
    for c in candidates:
        if c in headers:
            return c
        if c.lower() in header_lower:
            return header_lower[c.lower()]
    return None


def process_directory(batch_dir: Path) -> list[dict]:
    """Process all xlsx files in a directory."""
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl required. pip install openpyxl")
        sys.exit(1)

    xlsx_files = sorted(batch_dir.glob("*.xlsx"))
    if not xlsx_files:
        print(f"No xlsx files found in {batch_dir}")
        return []

    all_rows: list[dict] = []
    total_stats = {"files": 0, "total_read": 0, "kept": 0, "skipped": 0}

    for fpath in xlsx_files:
        fname = fpath.name
        if fname.startswith("~") or fname.startswith("."):
            continue

        meta = parse_filename(fname)
        if not meta:
            print(f"  SKIP unparseable filename: {fname}")
            continue

        batch_code = f"SENTIMENT-{meta['platform'].upper()}-{meta['country_iso']}-{datetime.now().strftime('%Y%m%d')}"

        wb = openpyxl.load_workbook(fpath, read_only=True)
        ws = wb[wb.sheetnames[0]]
        headers = [str(c.value or "") for c in next(ws.iter_rows(min_row=1, max_row=1))]
        col_map = {h: i for i, h in enumerate(headers)}

        text_col = _find_column(headers, TEXT_COLUMN_CANDIDATES)
        emotion_col = _find_column(headers, EMOTION_COLUMN_CANDIDATES)
        brand_col = _find_column(headers, BRAND_COLUMN_CANDIDATES)
        topic_col = _find_column(headers, TOPIC_COLUMN_CANDIDATES)
        url_col = _find_column(headers, URL_COLUMN_CANDIDATES)
        date_col = _find_column(headers, DATE_COLUMN_CANDIDATES)

        if not text_col:
            print(f"  SKIP {fname}: no text column found (tried: {TEXT_COLUMN_CANDIDATES})")
            wb.close()
            continue

        file_kept = 0
        file_total = 0
        for row in ws.iter_rows(min_row=2):
            vals = [c.value for c in row]
            file_total += 1

            text = str(vals[col_map[text_col]] or "").strip() if text_col in col_map else ""
            if not text or len(text) < 15:
                continue

            emotion = str(vals[col_map[emotion_col]] or "").strip() if emotion_col and emotion_col in col_map else ""
            post_type = ""
            if "帖子类型" in col_map:
                post_type = str(vals[col_map["帖子类型"]] or "").strip()

            total_stats["total_read"] += 1

            if not _is_negative(emotion, post_type, text):
                total_stats["skipped"] += 1
                continue

            brand_raw = str(vals[col_map[brand_col]] or "").strip() if brand_col and brand_col in col_map else ""
            brand = normalize_brand(brand_raw) if brand_raw else None

            topic = str(vals[col_map[topic_col]] or "").strip() if topic_col and topic_col in col_map else ""
            raw_url = str(vals[col_map[url_col]] or "").strip() if url_col and url_col in col_map else ""
            raw_date = str(vals[col_map[date_col]] or "").strip() if date_col and date_col in col_map else ""

            pain_cat = _infer_pain_category(text, topic)
            intensity = _infer_intensity(text)

            subcat = infer_pain_subcategory(meta["product_line"], text, topic)
            if subcat != "其他" and subcat in PAIN_SUBCATEGORY_PARENT:
                pain_cat = PAIN_SUBCATEGORY_PARENT[subcat]

            neg_theme = topic[:100] if topic else (subcat if subcat != "其他" else "")
            collect_date = raw_date[:10] if raw_date else datetime.now().strftime("%Y-%m-%d")

            row_dict = {
                "国家": meta["country"],
                "区域cluster": meta["cluster"],
                "产品品线": meta["product_line"],
                "平台类型": meta["platform_type"],
                "平台": meta["platform"],
                "画像编码": "",
                "画像名称": "",
                "生命周期": "",
                "痛点大类(功能/价格/体验/服务/安全)": pain_cat,
                "负面主题": neg_theme,
                "负面原文摘录(本地语言)": text[:500],
                "负面原文摘录(中文翻译)": "",
                "频次估算": "1",
                "负面强度(高/中/低)": intensity,
                "竞品关联品牌": brand or "",
                "对应运营建议": "",
                "来源URL": raw_url[:300] if raw_url else "",
                "采集日期": collect_date,
                "批次编码": batch_code,
                "优先级": "P2",
            }
            all_rows.append(row_dict)
            file_kept += 1

        wb.close()
        total_stats["files"] += 1
        total_stats["kept"] += file_kept
        print(f"  {fname}: {file_total} rows -> {file_kept} negative kept "
              f"({meta['country']}/{meta['product_line']}/{meta['platform']})")

    print(f"\nTotal: {total_stats['files']} files, "
          f"{total_stats['total_read']} read, "
          f"{total_stats['kept']} kept, "
          f"{total_stats['skipped']} skipped")
    return all_rows


def dedup_against_existing(new_rows: list[dict]) -> list[dict]:
    existing_keys: set[str] = set()
    if TARGET_CSV.exists():
        with open(TARGET_CSV, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for r in reader:
                url = (r.get("来源URL") or "").strip()
                text_prefix = (r.get("负面原文摘录(本地语言)") or "")[:50]
                existing_keys.add(f"{url}||{text_prefix}")

    deduped = []
    new_keys: set[str] = set()
    for r in new_rows:
        url = r.get("来源URL", "")
        text_prefix = r.get("负面原文摘录(本地语言)", "")[:50]
        key = f"{url}||{text_prefix}"
        if key in existing_keys or key in new_keys:
            continue
        new_keys.add(key)
        deduped.append(r)
    removed = len(new_rows) - len(deduped)
    print(f"Dedup: {len(new_rows)} -> {len(deduped)} (removed {removed})")
    return deduped


def append_to_csv(rows: list[dict]) -> None:
    file_exists = TARGET_CSV.exists() and TARGET_CSV.stat().st_size > 0
    with open(TARGET_CSV, "a", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_20_HEADERS)
        if not file_exists:
            writer.writeheader()
        writer.writerows(rows)
    print(f"Appended {len(rows)} rows to {TARGET_CSV.name}")


def save_staging(rows: list[dict], batch_dir_name: str) -> Path:
    staging_path = STAGING_DIR / f"sentiment_{batch_dir_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    staging_path.parent.mkdir(parents=True, exist_ok=True)

    country_dist: dict[str, int] = {}
    line_dist: dict[str, int] = {}
    platform_dist: dict[str, int] = {}
    brand_dist: dict[str, int] = {}
    for r in rows:
        c = r.get("国家", "?")
        country_dist[c] = country_dist.get(c, 0) + 1
        l = r.get("产品品线", "?")
        line_dist[l] = line_dist.get(l, 0) + 1
        p = r.get("平台", "?")
        platform_dist[p] = platform_dist.get(p, 0) + 1
        b = r.get("竞品关联品牌") or "未指定"
        brand_dist[b] = brand_dist.get(b, 0) + 1

    with open(staging_path, "w", encoding="utf-8") as f:
        json.dump({
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "source_dir": batch_dir_name,
            "total_rows": len(rows),
            "country_distribution": dict(sorted(country_dist.items(), key=lambda x: -x[1])),
            "product_line_distribution": dict(sorted(line_dist.items(), key=lambda x: -x[1])),
            "platform_distribution": dict(sorted(platform_dist.items(), key=lambda x: -x[1])),
            "brand_distribution": dict(sorted(brand_dist.items(), key=lambda x: -x[1])),
            "rows": rows,
        }, f, ensure_ascii=False, indent=2)
    print(f"Staging saved: {staging_path.name}")
    return staging_path


def main():
    dry_run = "--dry-run" in sys.argv
    batch_dir = DEFAULT_BATCH_DIR
    for i, arg in enumerate(sys.argv):
        if arg == "--dir" and i + 1 < len(sys.argv):
            batch_dir = Path(sys.argv[i + 1])

    print(f"=== Sentiment Batch Ingest ===")
    print(f"Source dir: {batch_dir}")
    print(f"Target CSV: {TARGET_CSV.name}")
    print(f"Mode:       {'DRY RUN' if dry_run else 'LIVE'}\n")

    if not batch_dir.exists():
        print(f"Directory not found: {batch_dir}")
        print(f"Create it and place xlsx files inside, using naming convention:")
        print(f"  {{platform}}_{{country_iso}}_{{product_line}}_sentiment_merged.xlsx")
        print(f"  e.g. reddit_us_breastpump_sentiment_merged.xlsx")
        return

    rows = process_directory(batch_dir)
    if not rows:
        print("No negative rows extracted.")
        return

    rows = dedup_against_existing(rows)
    if not rows:
        print("All rows are duplicates.")
        return

    staging = save_staging(rows, batch_dir.name)

    print(f"\n--- Distribution ---")
    dist_keys = ["国家", "产品品线", "平台", "竞品关联品牌"]
    for key in dist_keys:
        counts: dict[str, int] = {}
        for r in rows:
            v = r.get(key) or "未指定"
            counts[v] = counts.get(v, 0) + 1
        print(f"\n  {key}:")
        for k, v in sorted(counts.items(), key=lambda x: -x[1])[:10]:
            print(f"    {k}: {v}")

    if dry_run:
        print(f"\n[DRY RUN] Would append {len(rows)} rows. Skipping CSV write.")
    else:
        append_to_csv(rows)
        print(f"\nDone. {len(rows)} sentiment VOC records ingested.")
        print(f"Next steps:")
        print(f"  1. python tools/export_viz_json.py")
        print(f"  2. Reload backend: POST /api/v1/research/reload")


if __name__ == "__main__":
    main()
