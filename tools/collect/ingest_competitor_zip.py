# -*- coding: utf-8 -*-
"""
ingest_competitor_zip.py — 解压 01_raw_competitor.zip，
提取竞品差评 → 标准化 20 列追加到 dim_voc_negative_extract.csv。

流程:
  1. 解压 ZIP，跳过 __MACOSX
  2. 读取每个 xlsx 的 "统一分析文本"、"情绪"、"提及品牌" 等列
  3. 只保留 情绪=="差评" 或 (帖子类型=="吐槽" 且 情绪!="好评") 的行
  4. 从文件名提取主品牌 hint，与"提及品牌"列联合归因
  5. 映射到 dim_voc_negative_extract 标准 20 列
  6. 去重后追加到 CSV

用法:
    python tools/collect/ingest_competitor_zip.py [--dry-run]
"""
from __future__ import annotations

import csv
import io
import json
import re
import sys
import zipfile
from datetime import datetime, timezone
from pathlib import Path

PROJ = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(PROJ))

from tools.collect.brand_scope_config import REDDIT_FILENAME_BRAND_HINT

ZIP_PATH = PROJ / "data" / "add_data" / "01_raw_competitor.zip"
TARGET_CSV = PROJ / "data" / "delivery" / "tables" / "dim_voc_negative_extract.csv"
STAGING_JSON = PROJ / "data" / "processed" / "competitor_reddit_staging.json"
BATCH_CODE = f"COMP-REDDIT-{datetime.now().strftime('%Y%m%d')}"

BRAND_FROM_FILENAME: dict[str, str] = REDDIT_FILENAME_BRAND_HINT

CSV_20_HEADERS = [
    "国家", "区域cluster", "产品品线", "平台类型", "平台",
    "画像编码", "画像名称", "生命周期", "痛点大类(功能/价格/体验/服务/安全)",
    "负面主题", "负面原文摘录(本地语言)", "负面原文摘录(中文翻译)",
    "频次估算", "负面强度(高/中/低)", "竞品关联品牌", "对应运营建议",
    "来源URL", "采集日期", "批次编码", "优先级",
]

NEGATIVE_KEYWORDS = [
    "issue", "problem", "hate", "awful", "terrible", "horrible", "broken",
    "worst", "disappointed", "frustrat", "leak", "broke", "defect", "fail",
    "waste", "useless", "regret", "return", "refund", "pain", "hurt",
    "suck", "crap", "garbage", "annoying", "malfunction", "recall",
]


def _is_negative(emotion: str | None, post_type: str | None, text: str | None) -> bool:
    """Determine if a row is negative VOC."""
    e = (emotion or "").strip()
    pt = (post_type or "").strip()
    if e == "差评":
        return True
    if pt == "吐槽" and e != "好评":
        return True
    if e == "好评":
        return False
    if text:
        lower = text.lower()
        if sum(1 for kw in NEGATIVE_KEYWORDS if kw in lower) >= 2:
            return True
    return False


def _infer_pain_category(text: str, topics: str) -> str:
    """Infer pain category from text and topics."""
    combined = (text + " " + topics).lower()
    if any(w in combined for w in ["price", "expensive", "cost", "afford", "cheap", "money", "价格", "贵"]):
        return "价格"
    if any(w in combined for w in ["suction", "power", "motor", "function", "功能", "吸力", "电机"]):
        return "功能"
    if any(w in combined for w in ["service", "support", "customer", "warranty", "服务", "售后"]):
        return "服务"
    if any(w in combined for w in ["safe", "danger", "hazard", "recall", "安全", "危险"]):
        return "安全"
    return "体验"


def _infer_intensity(text: str) -> str:
    """Infer negative intensity from text."""
    lower = text.lower()
    strong = ["hate", "worst", "terrible", "horrible", "awful", "waste of money",
              "never again", "dangerous", "defect", "recall", "scam"]
    if any(w in lower for w in strong):
        return "高"
    mild = ["okay", "not great", "could be better", "minor", "slight"]
    if any(w in lower for w in mild):
        return "低"
    return "中"


def _resolve_brand(filename_brand: str, mentioned_brands: str | None) -> str:
    """Resolve the competitor brand from filename hint and content."""
    if filename_brand != "__compare__":
        return filename_brand
    if mentioned_brands:
        brands = [b.strip() for b in mentioned_brands.replace(",", ";").split(";") if b.strip()]
        if brands:
            return brands[0]
    return "多品牌对比"


def _build_reddit_url(subreddit: str | None, post_id: str | None) -> str:
    """Construct a Reddit URL from subreddit and post ID."""
    sub = subreddit or "unknown"
    pid = post_id or "unknown"
    return f"https://www.reddit.com/r/{sub}/comments/{pid}/"


def process_zip() -> list[dict]:
    """Process ZIP and return list of standardized 20-column rows."""
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl is required. Install with: pip install openpyxl")
        sys.exit(1)

    zf = zipfile.ZipFile(ZIP_PATH)
    all_rows: list[dict] = []
    stats = {"total_read": 0, "negative_kept": 0, "positive_skipped": 0, "neutral_skipped": 0}

    for info in zf.infolist():
        if info.is_dir() or "__MACOSX" in info.filename:
            continue
        fname = info.filename.split("/")[-1]
        if fname not in BRAND_FROM_FILENAME:
            print(f"  SKIP unknown file: {fname}")
            continue

        filename_brand = BRAND_FROM_FILENAME[fname]
        data = zf.read(info.filename)
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
        ws = wb[wb.sheetnames[0]]

        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        col = {h: i for i, h in enumerate(headers)}

        file_kept = 0
        file_total = 0
        for row in ws.iter_rows(min_row=2):
            vals = [c.value for c in row]
            file_total += 1

            text = str(vals[col.get("统一分析文本", -1)] or "").strip() if "统一分析文本" in col else ""
            if not text or len(text) < 20:
                continue

            emotion = str(vals[col.get("情绪", -1)] or "").strip() if "情绪" in col else ""
            post_type = str(vals[col.get("帖子类型", -1)] or "").strip() if "帖子类型" in col else ""

            stats["total_read"] += 1

            if not _is_negative(emotion, post_type, text):
                if emotion == "好评":
                    stats["positive_skipped"] += 1
                else:
                    stats["neutral_skipped"] += 1
                continue

            subreddit = str(vals[col.get("社区名称", -1)] or "") if "社区名称" in col else ""
            post_id = str(vals[col.get("帖子ID", -1)] or "") if "帖子ID" in col else ""
            mentioned_brands = str(vals[col.get("提及品牌", -1)] or "") if "提及品牌" in col else ""
            main_topic = str(vals[col.get("主要话题", -1)] or "") if "主要话题" in col else ""
            sub_topic = str(vals[col.get("次要话题", -1)] or "") if "次要话题" in col else ""
            crawl_time = str(vals[col.get("抓取时间", -1)] or "") if "抓取时间" in col else ""
            intent_tag = str(vals[col.get("意图标签", -1)] or "") if "意图标签" in col else ""

            brand = _resolve_brand(filename_brand, mentioned_brands)
            url = _build_reddit_url(subreddit, post_id)
            pain_cat = _infer_pain_category(text, main_topic + " " + sub_topic)
            intensity = _infer_intensity(text)
            neg_theme = sub_topic[:100] if sub_topic and sub_topic != "None" else (intent_tag[:100] if intent_tag and intent_tag != "None" else main_topic)
            collect_date = crawl_time[:10] if crawl_time else datetime.now().strftime("%Y-%m-%d")

            row_dict = {
                "国家": "美国",
                "区域cluster": "北美高购买力区",
                "产品品线": "吸奶器",
                "平台类型": "垂类社区类",
                "平台": "Reddit",
                "画像编码": "EP-MOM",
                "画像名称": "泵奶妈妈",
                "生命周期": "哺乳期",
                "痛点大类(功能/价格/体验/服务/安全)": pain_cat,
                "负面主题": neg_theme if neg_theme != "None" else "",
                "负面原文摘录(本地语言)": text[:500],
                "负面原文摘录(中文翻译)": "",
                "频次估算": "1",
                "负面强度(高/中/低)": intensity,
                "竞品关联品牌": brand,
                "对应运营建议": "",
                "来源URL": url,
                "采集日期": collect_date,
                "批次编码": BATCH_CODE,
                "优先级": "P2",
            }
            all_rows.append(row_dict)
            file_kept += 1

        wb.close()
        print(f"  {fname}: {file_total} total -> {file_kept} negative kept (brand={filename_brand})")
        stats["negative_kept"] += file_kept

    zf.close()
    print(f"\nSummary: read={stats['total_read']}, kept={stats['negative_kept']}, "
          f"positive_skipped={stats['positive_skipped']}, neutral_skipped={stats['neutral_skipped']}")
    return all_rows


def dedup_against_existing(new_rows: list[dict]) -> list[dict]:
    """Remove duplicates against existing CSV using URL+text prefix key."""
    existing_keys: set[str] = set()
    if TARGET_CSV.exists():
        with open(TARGET_CSV, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for r in reader:
                url = (r.get("来源URL") or "").strip()
                text_prefix = (r.get("负面原文摘录(本地语言)") or "")[:50]
                existing_keys.add(f"{url}||{text_prefix}")

    deduped = []
    new_keys: set[str] = set()
    for r in new_rows:
        url = r.get("来源URL", "")
        text_prefix = r.get("负面原文摘录(本地语言)", "")[:50]
        key = f"{url}||{text_prefix}"
        if key in existing_keys or key in new_keys:
            continue
        new_keys.add(key)
        deduped.append(r)
    print(f"Dedup: {len(new_rows)} -> {len(deduped)} (removed {len(new_rows) - len(deduped)} duplicates)")
    return deduped


def append_to_csv(rows: list[dict]) -> None:
    """Append rows to the target CSV."""
    file_exists = TARGET_CSV.exists() and TARGET_CSV.stat().st_size > 0
    with open(TARGET_CSV, "a", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_20_HEADERS)
        if not file_exists:
            writer.writeheader()
        writer.writerows(rows)
    print(f"Appended {len(rows)} rows to {TARGET_CSV}")


def save_staging(rows: list[dict]) -> None:
    """Save staging data for audit."""
    STAGING_JSON.parent.mkdir(parents=True, exist_ok=True)
    with open(STAGING_JSON, "w", encoding="utf-8") as f:
        json.dump({
            "batch_code": BATCH_CODE,
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "total_rows": len(rows),
            "brand_distribution": _brand_dist(rows),
            "rows": rows,
        }, f, ensure_ascii=False, indent=2)
    print(f"Staging saved to {STAGING_JSON}")


def _brand_dist(rows: list[dict]) -> dict[str, int]:
    dist: dict[str, int] = {}
    for r in rows:
        b = r.get("竞品关联品牌", "unknown")
        dist[b] = dist.get(b, 0) + 1
    return dict(sorted(dist.items(), key=lambda x: -x[1]))


def main():
    dry_run = "--dry-run" in sys.argv

    print(f"=== Competitor ZIP Ingest ===")
    print(f"Source: {ZIP_PATH}")
    print(f"Target: {TARGET_CSV}")
    print(f"Batch:  {BATCH_CODE}")
    print(f"Mode:   {'DRY RUN' if dry_run else 'LIVE'}\n")

    if not ZIP_PATH.exists():
        print(f"ERROR: ZIP not found at {ZIP_PATH}")
        sys.exit(1)

    rows = process_zip()
    if not rows:
        print("No negative rows extracted.")
        return

    rows = dedup_against_existing(rows)
    if not rows:
        print("All rows are duplicates of existing data.")
        return

    save_staging(rows)

    print(f"\nBrand distribution:")
    for brand, count in _brand_dist(rows).items():
        print(f"  {brand}: {count}")

    if dry_run:
        print(f"\n[DRY RUN] Would append {len(rows)} rows. Skipping CSV write.")
    else:
        append_to_csv(rows)
        print(f"\nDone. Total {len(rows)} competitor negative reviews ingested.")


if __name__ == "__main__":
    main()
