from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook


HTML_PATTERN = re.compile(r"<[^>]+>|&[a-zA-Z0-9#]+;")


class WorkbookQaRunner:
    def __init__(self, product_root: Path) -> None:
        self.product_root = product_root
        self.target_dir = product_root / "data" / "delivery"

    def scan_html_pollution(self, max_cells: int = 20000) -> dict:
        issues: list[dict[str, str]] = []
        scanned = 0
        workbooks = list(self.target_dir.glob("*.xlsx"))

        for wb_path in workbooks:
            wb = load_workbook(wb_path, read_only=True, data_only=True)
            try:
                for ws in wb.worksheets:
                    for row in ws.iter_rows(values_only=True):
                        for cell_value in row:
                            if scanned >= max_cells:
                                break
                            scanned += 1
                            if isinstance(cell_value, str) and HTML_PATTERN.search(cell_value):
                                issues.append(
                                    {
                                        "workbook": wb_path.name,
                                        "sheet": ws.title,
                                        "sample": cell_value[:200],
                                    }
                                )
                        if scanned >= max_cells:
                            break
                    if scanned >= max_cells:
                        break
            finally:
                wb.close()

        return {
            "checked_workbooks": [p.name for p in workbooks],
            "scanned_cells": scanned,
            "issue_count": len(issues),
            "issues": issues[:100],
            "passed": len(issues) == 0,
        }

