"""Reader for research dimension tables (拆分后 CSV 维表 / 配置表)."""
from __future__ import annotations

import csv
from pathlib import Path


class ResearchReader:
    def __init__(self, product_root: Path) -> None:
        self.tables_dir = product_root / "data" / "delivery" / "tables"
        self._inputs_dir = product_root / "data" / "inputs"

    def _read_table(self, csv_name: str) -> list[dict[str, str]]:
        path = self.tables_dir / csv_name
        if not path.exists():
            return []
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            return list(csv.DictReader(f))

    @staticmethod
    def _filter(
        rows: list[dict[str, str]],
        filters: dict[str, str | None],
    ) -> list[dict[str, str]]:
        result = rows
        for key, value in filters.items():
            if value is not None:
                result = [
                    r for r in result
                    if (r.get(key) or "").strip().lower() == value.strip().lower()
                ]
        return result

    def get_persona(
        self,
        country: str | None = None,
        product_line: str | None = None,
    ) -> list[dict[str, str]]:
        rows = self._read_table("dim_country_product_persona.csv")
        return self._filter(rows, {"国家": country, "产品品线": product_line})

    def get_segments(
        self,
        country: str | None = None,
        product_line: str | None = None,
    ) -> list[dict[str, str]]:
        rows = self._read_table("dim_country_segment_matrix.csv")
        return self._filter(rows, {"国家": country, "产品品线": product_line})

    def get_price_sensitivity(
        self,
        country: str | None = None,
        product_line: str | None = None,
    ) -> list[dict[str, str]]:
        rows = self._read_table("dim_country_price_sensitivity.csv")
        return self._filter(rows, {"国家": country, "产品品线": product_line})

    def get_info_sources(
        self,
        country: str | None = None,
        product_line: str | None = None,
    ) -> list[dict[str, str]]:
        rows = self._read_table("dim_info_source_quality.csv")
        return self._filter(rows, {"国家": country, "产品品线": product_line})

    def get_cluster_strategy(self) -> list[dict[str, str]]:
        return self._read_table("dim_cluster_strategy.csv")

    def get_platform_entry(
        self,
        country: str | None = None,
    ) -> list[dict[str, str]]:
        rows = self._read_table("cfg_top10_platform_entry.csv")
        return self._filter(rows, {"国家": country})

    def get_country_insight(self) -> list[dict[str, str]]:
        return self._read_table("dim_top20_country_insight.csv")

    def get_search_playbook(
        self,
        country: str | None = None,
        product_line: str | None = None,
    ) -> list[dict[str, str]]:
        rows = self._read_table("cfg_p1_search_playbook.csv")
        return self._filter(rows, {"国家": country, "产品品线": product_line})

    def get_project_meta(self) -> list[dict[str, str]]:
        return self._read_table("dim_project_meta.csv")

    def get_top10_country_line(
        self,
        country: str | None = None,
    ) -> list[dict[str, str]]:
        rows = self._read_table("cfg_top10_country_line.csv")
        return self._filter(rows, {"国家": country})

    def get_input_scope(self) -> list[dict[str, str]]:
        """Read the original input Excel (国家-品线找用户集中地.xlsx)."""
        xlsx_path = self._inputs_dir / "国家-品线找用户集中地.xlsx"
        if not xlsx_path.exists():
            return []
        try:
            import openpyxl
        except ImportError:
            return []
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(h or "").strip() for h in next(rows_iter, ())]
        if not headers:
            wb.close()
            return []
        result = []
        for vals in rows_iter:
            row = {headers[i]: str(v if v is not None else "") for i, v in enumerate(vals) if i < len(headers)}
            result.append(row)
        wb.close()
        return result

    def available_tables(self) -> list[dict[str, str | int]]:
        """List all CSV tables with row counts."""
        results: list[dict[str, str | int]] = []
        if not self.tables_dir.exists():
            return results
        for csv_file in sorted(self.tables_dir.glob("*.csv")):
            with csv_file.open("r", encoding="utf-8-sig", newline="") as f:
                reader = csv.reader(f)
                count = sum(1 for _ in reader) - 1
            results.append({"table": csv_file.name, "rows": max(0, count)})
        return results
